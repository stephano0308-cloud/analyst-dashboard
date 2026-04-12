from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


def docs_dir() -> Path:
    return Path(os.getenv("MCP_DOCS_DIR", Path(__file__).resolve().parents[1] / "data" / "docs"))


def list_documents() -> list[dict[str, Any]]:
    base = docs_dir()
    if not base.exists():
        return []

    rows: list[dict[str, Any]] = []
    for path in sorted(base.rglob("*")):
        if path.is_file():
            rows.append({
                "path": str(path),
                "name": path.name,
                "suffix": path.suffix.lower(),
                "size": path.stat().st_size,
            })
    return rows


def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def _read_json(path: Path) -> str:
    return json.dumps(json.loads(_read_text(path)), ensure_ascii=False, indent=2)


def _read_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    texts = []
    for page in reader.pages:
        texts.append(page.extract_text() or "")
    return "\n\n".join(texts)


def _read_excel(path: Path) -> str:
    wb = load_workbook(filename=str(path), data_only=True)
    parts = []
    for ws in wb.worksheets:
        parts.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            line = " | ".join("" if v is None else str(v) for v in row)
            if line.strip(" |"):
                parts.append(line)
    return "\n".join(parts)


def read_document(path: str) -> dict[str, Any]:
    p = Path(path)
    if not p.exists():
        return {"path": path, "status": "missing", "text": ""}

    suffix = p.suffix.lower()
    if suffix in {".txt", ".md", ".csv"}:
        text = _read_text(p)
    elif suffix == ".json":
        text = _read_json(p)
    elif suffix == ".pdf":
        text = _read_pdf(p)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        text = _read_excel(p)
    else:
        text = _read_text(p)

    return {
        "path": str(p),
        "status": "ok",
        "type": suffix.lstrip("."),
        "text": text,
    }
